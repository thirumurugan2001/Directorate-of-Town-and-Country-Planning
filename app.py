import os
import sys
import pandas as pd
from urllib.parse import urljoin
from playwright.sync_api import sync_playwright
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox,
    QPushButton, QTextEdit, QFileDialog
)
from PyQt5.QtCore import QThread, pyqtSignal
from cropper import process_pdf_from_url                
from model import extract_pdf_details_from_image    


def normalize_signature(sig):
    if isinstance(sig, dict):
        return {
            "name_Address": sig.get("name_Address", "") if sig.get("name_Address", "") is not None else "",
            "mail": sig.get("mail", "") if sig.get("mail", "") is not None else "",
            "phone": sig.get("phone", "") if sig.get("phone", "") is not None else ""
        }
    if isinstance(sig, str) and sig.strip():
        return {"name_Address": sig.strip(), "mail": "", "phone": ""}
    return {"name_Address": "", "mail": "", "phone": ""}

# ---------------- Worker Thread ----------------
class ScraperThread(QThread):
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(str)
    def __init__(self, department, approval_type, district, year, output_excel):
        super().__init__()
        self.department = department
        self.approval_type = approval_type
        self.district = district
        self.year = year
        self.output_excel = output_excel
        self.output_dir = "cropped_images"
        os.makedirs(self.output_dir, exist_ok=True)
    def log(self, message: str):
        self.log_signal.emit(message)
        
    def run(self):
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            page.goto("https://onlineppa.tn.gov.in/approved-plan-list", timeout=60000)
            self.log("Selecting dropdown values...")
            page.evaluate(f"""
                () => {{
                    function selectNiceOption(selectId, visibleText) {{
                        const selectEl = document.querySelector('select#' + selectId);
                        const niceDiv = selectEl?.nextElementSibling;
                        if (!selectEl || !niceDiv) return;

                        const options = niceDiv.querySelectorAll('li.option');
                        const realOption = Array.from(options).find(li => li.textContent.trim() === visibleText);
                        if (realOption) {{
                            const current = niceDiv.querySelector('.current');
                            if (current) current.textContent = realOption.textContent;
                            realOption.classList.add('selected');
                            selectEl.value = realOption.getAttribute('data-value');
                            selectEl.dispatchEvent(new Event('change', {{ bubbles: true }}));
                        }}
                    }}
                    selectNiceOption('depName', '{self.department}');
                    selectNiceOption('appType', '{self.approval_type}');
                    selectNiceOption('district', '{self.district}');
                    selectNiceOption('year', '{self.year}');
                }}
            """)
            page.wait_for_timeout(2000)
            page.evaluate("document.querySelector('#search')?.click()")
            page.wait_for_function("""
                () => {
                    const tbody = document.querySelector("table.PPAData tbody");
                    return tbody && tbody.querySelectorAll("tr").length > 0;
                }
            """, timeout=120000)
            self.log("Scraping results...")
            rows_data = []
            scraped_count = 0
            page_num = 1
            while scraped_count < 5:
                self.log(f"Scraping page {page_num}...")
                rows = page.query_selector_all("table.PPAData tbody tr")
                for row in rows:
                    if scraped_count >= 5:
                        break
                    cells = row.query_selector_all("td")
                    if len(cells) < 10:
                        continue
                    application_no     = cells[1].inner_text().strip()
                    district_val       = cells[2].inner_text().strip()
                    approval_type_val  = cells[3].inner_text().strip()
                    permit_date        = cells[4].inner_text().strip()
                    total_fees         = cells[5].inner_text().strip()
                    project_title        = "N/A"
                    applicant_signature  = ""                    
                    architect_signature  = {"name_Address": "", "mail": "", "phone": ""}
                    pdf_url_for_link     = ""
                    try:
                        pdf_element = cells[9].query_selector("a")
                        if pdf_element:
                            pdf_url = pdf_element.get_attribute("href") or ""
                            pdf_url_for_link = urljoin(page.url, pdf_url)
                            print("PDF URL:", pdf_url_for_link)
                            img, msg = process_pdf_from_url(pdf_url_for_link)
                            self.log(f"{application_no}: {msg}")
                            if img:
                                filename = f"{application_no}_full.png".replace("/", "_")
                                filepath = os.path.join(self.output_dir, filename)
                                img.save(filepath)
                                try:
                                    seal_data = extract_pdf_details_from_image(filepath) or {}
                                except Exception as me:
                                    self.log(f"{application_no}: error calling model - {me}")
                                    seal_data = {}
                                project_title       = seal_data.get("PROJECT TITLE", "N/A")
                                applicant_signature = seal_data.get("OWNER SIGNATURE", "")                                
                                architect_signature  = normalize_signature(seal_data.get("REGISTERED ENGINEER", {}))
                    except Exception as e:
                        self.log(f"{application_no}: Error processing PDF/model - {e}")
                    rows_data.append({
                        "Application No": application_no,
                        "District": district_val,
                        "Approval Type": approval_type_val,
                        "Permit Issue Date": permit_date,
                        "Total Fees": total_fees,
                        "Project Title": project_title,
                        "Applicant/Owner Signature": applicant_signature,                        
                        "Registered Engineer Name/Address": architect_signature.get("name_Address", ""),
                        "Registered Engineer Mail": architect_signature.get("mail", ""),
                        "Registered Engineer Phone": architect_signature.get("phone", ""),
                        "PDF URL": pdf_url_for_link,
                        "PDF Link": "View PDF"
                    })
                next_btn_li = page.query_selector("li#example_next")
                if next_btn_li and "disabled" not in (next_btn_li.get_attribute("class") or ""):
                    next_btn_li.click()
                    page.wait_for_timeout(2000)
                    page_num += 1
                else:
                    break
            browser.close()
        df = pd.DataFrame(rows_data, columns=[
            "Application No", "District", "Approval Type",
            "Permit Issue Date", "Total Fees",
            "Project Title", "Applicant/Owner Signature",
            "Architect Name/Address", "Architect Mail", "Architect Phone",
            "PDF URL"
        ])
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            headers = [
                "Application No", "District", "Approval Type",
                "Permit Issue Date", "Total Fees",
                "Project Title", "Applicant/Owner Signature",
                "Registered Engineer Name/Address", "Registered Engineer Mail", "Registered Engineer Phone",
                "PDF Link"
            ]
            ws.append(headers)
            for row in rows_data:
                ws.append([
                    row["Application No"],
                    row["District"],
                    row["Approval Type"],
                    row["Permit Issue Date"],
                    row["Total Fees"],
                    row["Project Title"],
                    row["Applicant/Owner Signature"],
                    row["Registered Engineer Name/Address"],
                    row["Registered Engineer Mail"],
                    row["Registered Engineer Phone"],
                    "View PDF"])
                cell = ws.cell(row=ws.max_row, column=11)
                cell.hyperlink = row.get("PDF URL", "")
                cell.font = Font(color="0000FF", underline="single")
            wb.save(self.output_excel)
            self.finished_signal.emit(f"Excel saved to: {self.output_excel}")
        except Exception as e:
            self.finished_signal.emit(f"Failed to write Excel: {e}")

class ScraperApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SWP PDF Scraper")
        self.setGeometry(200, 200, 700, 420)
        layout = QVBoxLayout()
        self.department = QComboBox()
        self.department.addItems(["DTCP", "CMDA"])
        self.approval_type = QComboBox()
        self.approval_type.addItems(["Building Plan", "Layout Plan"])
        self.district = QComboBox()
        self.district.addItems(["Chengalpattu", "Chennai", "Coimbatore", "Kallakurichi", "Kancheepuram", "Mayiladuthurai",
                                "Perambalur","Ranipet", "Tiruvannamalai", "Tiruvarur", "Vellore", "Viluppuram", "Virudhunagar"])
        self.year = QComboBox()
        self.year.addItems(["2023", "2024", "2025"])
        form_layout = QHBoxLayout()
        form_layout.addWidget(QLabel("Department:"))
        form_layout.addWidget(self.department)
        form_layout.addWidget(QLabel("Approval Type:"))
        form_layout.addWidget(self.approval_type)
        form_layout.addWidget(QLabel("District:"))
        form_layout.addWidget(self.district)
        form_layout.addWidget(QLabel("Year:"))
        form_layout.addWidget(self.year)
        layout.addLayout(form_layout)
        self.go_button = QPushButton("Go")
        self.go_button.clicked.connect(self.start_scraping)
        layout.addWidget(self.go_button)
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        layout.addWidget(self.log_box)
        self.setLayout(layout)

    def log_message(self, msg: str):
        self.log_box.append(msg)

    def start_scraping(self):
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if not save_path:
            return
        self.log_box.clear()
        self.log_message("Starting scraper...")
        self.thread = ScraperThread(
            self.department.currentText(),
            self.approval_type.currentText(),
            self.district.currentText(),
            self.year.currentText(),
            save_path
        )
        self.thread.log_signal.connect(self.log_message)
        self.thread.finished_signal.connect(self.log_message)
        self.thread.start()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = ScraperApp()
    win.show()
    sys.exit(app.exec_())
