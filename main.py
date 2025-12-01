import os
import sys
import openpyxl
import pandas as pd
from style import AppStyle
from pathlib import Path
from PyQt5.QtCore import Qt
from openpyxl import Workbook
from urllib.parse import urljoin
from openpyxl.styles import Font
from cropper import process_pdf_from_url                
from PyQt5.QtCore import QThread, pyqtSignal
from playwright.sync_api import sync_playwright
from model import extract_pdf_details_from_image    
from helper import normalize_signature, extract_contact_from_text
from PyQt5.QtGui import QFont, QColor, QPalette, QPixmap, QMovie, QIcon
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton, QFileDialog, QGroupBox, QProgressBar, QMessageBox,QSpacerItem, QSizePolicy)

class ScraperThread(QThread):
    finished_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int, int)
    
    def __init__(self, department, approval_type, district, year, output_excel):
        super().__init__()
        self.department = department
        self.approval_type = approval_type
        self.district = district
        self.year = year
        self.output_excel = output_excel
        self.output_dir = "cropped_images"
        os.makedirs(self.output_dir, exist_ok=True)

    def run(self):
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=False)
                page = browser.new_page()
                page.goto("https://onlineppa.tn.gov.in/approved-plan-list", timeout=100000)                
                page.evaluate(f"""
                    () => {{
                        function selectNiceOption(selectId, visibleText) {{
                            const selectEl = document.querySelector('select#' + selectId);
                            const niceDiv = selectEl?.nextElementSibling;
                            if (!selectEl || !niceDiv) return;

                            const options = niceDiv.querySelectorAll('li.option');
                            const realOption = Array.from(options).find(li => li.textContent.trim() === visibleText);
                            if (realOption) {{
                                const current = niceDiv.querySelector('.current');
                                if (current) current.textContent = realOption.textContent;
                                realOption.classList.add('selected');
                                selectEl.value = realOption.getAttribute('data-value');
                                selectEl.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            }}
                        }}
                        selectNiceOption('depName', '{self.department}');
                        selectNiceOption('appType', '{self.approval_type}');
                        selectNiceOption('district', '{self.district}');
                        selectNiceOption('year', '{self.year}');
                    }}
                """)
                page.wait_for_timeout(2000)
                page.evaluate("document.querySelector('#search')?.click()")                
                page.wait_for_function("""
                    () => {
                        const tbody = document.querySelector("table.PPAData tbody");
                        return tbody && tbody.querySelectorAll("tr").length > 0;
                    }
                """, timeout=120000)                
                rows_data = []
                scraped_count = 0
                page_num = 1
                total_rows = 0                
                try:
                    total_text = page.inner_text("div.dataTables_info")
                    import re
                    match = re.search(r'of\s+(\d+)', total_text)
                    if match:
                        total_rows = int(match.group(1))
                    else:
                        total_rows = 50  
                except:
                    total_rows = 50                
                while scraped_count < 5: 
                    rows = page.query_selector_all("table.PPAData tbody tr")                    
                    for i, row in enumerate(rows):
                        if scraped_count >= 5:
                            break                            
                        current_count = scraped_count + 1
                        self.progress_signal.emit(current_count, min(10, total_rows))                        
                        cells = row.query_selector_all("td")
                        if len(cells) < 10:
                            continue                            
                        s_no = cells[0].inner_text().strip()
                        application_no = cells[1].inner_text().strip()
                        district_val = cells[2].inner_text().strip()
                        approval_type_val = cells[3].inner_text().strip()
                        permit_date = cells[4].inner_text().strip()
                        total_fees = cells[5].inner_text().strip()
                        approved_plan_url = ""
                        project_title = "N/A"
                        applicant_signature = ""                    
                        architect_signature = {"name_Address": "", "mail": "", "phone": ""}
                        structural_engineer_signature = {"name_Address": "", "mail": "", "phone": ""}
                        
                        try:
                            approved_plan_element = cells[9].query_selector("a")
                            if approved_plan_element:
                                approved_plan_href = approved_plan_element.get_attribute("href") or ""
                                approved_plan_url = urljoin(page.url, approved_plan_href)                                
                                img, msg = process_pdf_from_url(approved_plan_url)                                
                                if img:
                                    filename = f"{application_no}_full.png".replace("/", "_")
                                    filepath = os.path.join(self.output_dir, filename)
                                    img.save(filepath)                                    
                                    try:
                                        seal_data = extract_pdf_details_from_image(filepath) or {}
                                        print(f"[DEBUG] Model response for {application_no}: {seal_data}")
                                    except Exception as me:
                                        print(f"[ERROR] Model extraction failed: {me}")
                                        seal_data = {}                                    
                                    if isinstance(seal_data, dict) and seal_data.get("status"):
                                        project_title = seal_data.get("PROJECT TITLE", "N/A")
                                        applicant_signature = seal_data.get("OWNER SIGNATURE", "")                                        
                                        architect_info = seal_data.get("REGISTERED ENGINEER", "")
                                        if isinstance(architect_info, dict):
                                            architect_signature = normalize_signature(architect_info)
                                        elif isinstance(architect_info, str):
                                            email, phone = extract_contact_from_text(architect_info)
                                            architect_signature = {
                                                "name_Address": architect_info.strip(),
                                                "mail": email,
                                                "phone": phone
                                            }                                        
                                        structural_info = seal_data.get("STRUCTURAL ENGINEER", "")
                                        if isinstance(structural_info, dict):
                                            structural_engineer_signature = normalize_signature(structural_info)
                                        elif isinstance(structural_info, str):
                                            # Try to extract email and phone from the string
                                            email, phone = extract_contact_from_text(structural_info)
                                            structural_engineer_signature = {
                                                "name_Address": structural_info.strip(),
                                                "mail": email,
                                                "phone": phone
                                            }
                                    else:
                                        if isinstance(seal_data, dict):
                                            project_title = seal_data.get("PROJECT TITLE", "N/A")
                                            applicant_signature = seal_data.get("OWNER SIGNATURE", "")                                            
                                            structural_info = seal_data.get("STRUCTURAL ENGINEER", "") or seal_data.get("data", {}).get("STRUCTURAL ENGINEER", "")
                                            if structural_info:
                                                email, phone = extract_contact_from_text(str(structural_info))
                                                structural_engineer_signature = {
                                                    "name_Address": str(structural_info).strip(),
                                                    "mail": email,
                                                    "phone": phone
                                                }
                        except Exception as e:
                            print(f"[ERROR] PDF processing failed: {e}")
                            pass                   
                        demand_details_url = ""
                        try:
                            demand_details_element = cells[10].query_selector("a")
                            if demand_details_element:
                                demand_details_href = demand_details_element.get_attribute("href") or ""
                                demand_details_url = urljoin(page.url, demand_details_href)
                        except:
                            pass                        
                        rows_data.append({
                            "S.No": s_no,
                            "Application No": application_no,
                            "District": district_val,
                            "Approval Type": approval_type_val,
                            "Permit Issue Date": permit_date,
                            "Total Fees": total_fees,
                            "Approved Plan URL": approved_plan_url,
                            "Demand Details URL": demand_details_url,
                            "Project Title": project_title,
                            "Applicant/Owner Signature": applicant_signature,                        
                            "Registered Engineer Name/Address": architect_signature.get("name_Address", ""),
                            "Registered Engineer Mail": architect_signature.get("mail", ""),
                            "Registered Engineer Phone": architect_signature.get("phone", ""),
                            "Structural Engineer Name/Address": structural_engineer_signature.get("name_Address", ""),
                            "Structural Engineer Mail": structural_engineer_signature.get("mail", ""),
                            "Structural Engineer Phone": structural_engineer_signature.get("phone", ""),
                            "Approved Plan Link": "Download" if approved_plan_url else "N/A",
                            "Demand Details Link": "Download" if demand_details_url else "N/A"
                        })                        
                        scraped_count += 1                    
                    next_btn_li = page.query_selector("li#example_next")
                    if next_btn_li and "disabled" not in (next_btn_li.get_attribute("class") or ""):
                        next_btn_li.click()
                        page.wait_for_timeout(2000)
                        page_num += 1
                    else:
                        break                        
                browser.close()                
            self.save_to_excel(rows_data)
            self.finished_signal.emit(f"Excel saved to: {self.output_excel}")            
        except Exception as e:
            self.finished_signal.emit(f"Scraping failed: {e}")

    def save_to_excel(self, rows_data):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "DTCP Results"            
            headers = [
                "S.No", "Application No", "District", "Approval Type",
                "Permit Issue Date", "Total Fees", "Approved Plan", "Demand Details",
                "Project Title", "Applicant/Owner Signature",
                "Registered Engineer Name/Address", "Registered Engineer Mail", "Registered Engineer Phone",
                "Structural Engineer Name/Address", "Structural Engineer Mail", "Structural Engineer Phone"
            ]
            ws.append(headers)          
            column_widths = {
                'A': 8,    
                'B': 20,  
                'C': 15,  
                'D': 18,   
                'E': 15,   
                'F': 15,  
                'G': 15,  
                'H': 15,   
                'I': 40,  
                'J': 30,   
                'K': 40,  
                'L': 25,  
                'M': 20,   
                'N': 40,  
                'O': 25,  
                'P': 20,
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width            
            for row in rows_data:
                ws.append([
                    row["S.No"],
                    row["Application No"],
                    row["District"],
                    row["Approval Type"],
                    row["Permit Issue Date"],
                    row["Total Fees"],
                    "Download" if row["Approved Plan URL"] else "N/A",
                    "Download" if row["Demand Details URL"] else "N/A",
                    row["Project Title"],
                    row["Applicant/Owner Signature"],
                    row["Registered Engineer Name/Address"],
                    row["Registered Engineer Mail"],
                    row["Registered Engineer Phone"],
                    row["Structural Engineer Name/Address"],  
                    row["Structural Engineer Mail"],  
                    row["Structural Engineer Phone"]
                ])
                approved_plan_cell = ws.cell(row=ws.max_row, column=7)
                if row["Approved Plan URL"]:
                    approved_plan_cell.hyperlink = row["Approved Plan URL"]
                    approved_plan_cell.font = Font(color="0000FF", underline="single")                
                demand_details_cell = ws.cell(row=ws.max_row, column=8)
                if row["Demand Details URL"]:
                    demand_details_cell.hyperlink = row["Demand Details URL"]
                    demand_details_cell.font = Font(color="0000FF", underline="single")                
                wrap_columns = [9, 10, 11, 14] 
                for col in wrap_columns:
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)                
            wb.save(self.output_excel)            
        except Exception as e:
            raise Exception(f"Failed to write Excel: {e}")

class DTCPApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DTCP Plan Permit Scraper - Ajantha Bathroom Products")
        self.setMinimumSize(1200, 800)        
        self.set_window_icon()        
        self.setup_ui()    
    
    def set_window_icon(self):
        icon_path = str(Path(__file__).resolve().parent / "client_logo.png")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        else:
            pixmap = QPixmap(32, 32)
            pixmap.fill(QColor("#8f0606"))
            self.setWindowIcon(QIcon(pixmap))
    
    def setup_ui(self):
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(30, 20, 30, 20)
        self.layout.setSpacing(15)        
        palette = self.palette()
        palette.setColor(QPalette.Window, QColor("#000000"))
        palette.setColor(QPalette.WindowText, QColor("#ffffff"))
        palette.setColor(QPalette.Base, QColor("#1a1a1a"))
        palette.setColor(QPalette.AlternateBase, QColor("#2a2a2a"))
        palette.setColor(QPalette.Text, QColor("#ffffff"))
        palette.setColor(QPalette.Button, QColor("#dc2626"))
        palette.setColor(QPalette.ButtonText, QColor("#ffffff"))
        self.setPalette(palette)
        self.setAutoFillBackground(True)        
        self.setStyleSheet(AppStyle)       
        header_layout = QHBoxLayout()
        header_layout.setSpacing(30)
        header_layout.setAlignment(Qt.AlignCenter)
        self.logo_label = QLabel()
        logo_path = str(Path(__file__).resolve().parent / "client_logo.png")
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            scaled_pixmap = pixmap.scaled(120, 120, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.logo_label.setPixmap(scaled_pixmap)
        else:
            self.logo_label.setText("🏭")
            self.logo_label.setStyleSheet("font-size: 80px; color: #dc2626;")
        self.logo_label.setAlignment(Qt.AlignCenter)        
        text_layout = QVBoxLayout()
        text_layout.setSpacing(5)
        text_layout.setAlignment(Qt.AlignCenter)        
        company_label = QLabel("AJANTHA BATHROOM PRODUCTS")
        company_label.setObjectName("CompanyName")
        company_label.setAlignment(Qt.AlignCenter)
        company_label.setFont(QFont("Segoe UI", 20, QFont.Black))        
        product_label = QLabel("AND PIPES PRIVATE LIMITED")
        product_label.setObjectName("ProductName")
        product_label.setAlignment(Qt.AlignCenter)
        product_label.setFont(QFont("Segoe UI", 16, QFont.Bold))        
        desc_label = QLabel("Premium Bathroom Solutions & Sanitaryware Manufacturers")
        desc_label.setObjectName("ProductDesc")
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setFont(QFont("Segoe UI", 12, QFont.Normal))        
        text_layout.addWidget(company_label)
        text_layout.addWidget(product_label)
        text_layout.addWidget(desc_label)
        header_layout.addWidget(self.logo_label)
        header_layout.addLayout(text_layout)        
        self.layout.addLayout(header_layout)        
        separator = QLabel()
        separator.setFixedHeight(2)
        separator.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #dc2626, stop:0.5 #ffffff, stop:1 #dc2626); margin: 15px 0px;")
        self.layout.addWidget(separator)        
        main_title = QLabel("DTCP PLAN PERMIT SCRAPER")
        main_title.setObjectName("MainTitle")
        main_title.setAlignment(Qt.AlignCenter)
        main_title.setFont(QFont("Segoe UI", 24, QFont.Black))
        self.layout.addWidget(main_title)        
        subtitle = QLabel("Professional Data Extraction & Analysis System")
        subtitle.setObjectName("SubTitleLabel")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setFont(QFont("Segoe UI", 12))
        self.layout.addWidget(subtitle)        
        output_group = QGroupBox("Application Output")
        output_layout = QVBoxLayout()        
        search_group = QGroupBox("Search Configuration")
        search_layout = QVBoxLayout()        
        filters_layout = QHBoxLayout()
        filters_layout.setAlignment(Qt.AlignCenter)        
        dept_layout = QVBoxLayout()
        dept_layout.addWidget(QLabel("Department Name"))
        self.department = QComboBox()
        self.department.addItems(["DTCP", "Rural Panchayat", "Town Panchayat"])
        self.department.setFixedWidth(200)
        dept_layout.addWidget(self.department)
        filters_layout.addLayout(dept_layout)        
        type_layout = QVBoxLayout()
        type_layout.addWidget(QLabel("Approval Type"))
        self.approval_type = QComboBox()
        self.approval_type.addItems(["Layout Approval", "Building Plan"])
        self.approval_type.setFixedWidth(200)
        type_layout.addWidget(self.approval_type)
        filters_layout.addLayout(type_layout)        
        district_layout = QVBoxLayout()
        district_layout.addWidget(QLabel("District"))
        self.district = QComboBox()
        districts = ["Ariyalur", "Chengalpattu", "Chennai", "Coimbatore", "Cuddalore","Dharmapuri", "Dindigul", "Erode", "Kallakurichi", "Kancheepuram","Kanniyakumari", "Karur", "Krishnagiri", "Madurai", "Mayiladuthurai","Nagapattinam", "Namakkal", "Perambalur", "Pudukkottai", "Ramanathapuram","Ranipet", "Salem", "Sivagangai", "Tenkasi", "Thanjavur","Theni", "The Nilgiris", "Thoothukkudi", "Tiruchirappalli", "Tirunelveli","Tirupathur", "Tiruppur", "Tiruvallur", "Tiruvannamalai", "Tiruvarur","Vellore", "Vilupuram", "Virudhunagar"]
        self.district.addItems(districts)
        self.district.setFixedWidth(200)
        district_layout.addWidget(self.district)
        filters_layout.addLayout(district_layout)        
        year_layout = QVBoxLayout()
        year_layout.addWidget(QLabel("Year"))
        self.year = QComboBox()
        self.year.addItems(["2022", "2023", "2024", "2025"])
        self.year.setFixedWidth(150)
        year_layout.addWidget(self.year)
        filters_layout.addLayout(year_layout)        
        search_layout.addLayout(filters_layout)
        search_group.setLayout(search_layout)
        output_layout.addWidget(search_group)        
        progress_group = QGroupBox("Extraction Progress")
        progress_layout = QVBoxLayout()        
        self.loader = QLabel()
        self.loader.setAlignment(Qt.AlignCenter)
        self.loader.setMinimumHeight(80)
        gif_path = str(Path(__file__).resolve().parent / "loader.gif")
        if os.path.exists(gif_path):
            self.loader_movie = QMovie(gif_path)
            self.loader.setMovie(self.loader_movie)
        else:
            self.loader.setText("Processing...")
            self.loader.setStyleSheet("font-size: 16px; color: #dc2626; font-weight: bold;")
        self.loader.setVisible(False)
        progress_layout.addWidget(self.loader)        
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.progress.setFormat("Processing: %p% (%v of %m documents)")
        self.progress.setFixedHeight(30)
        progress_layout.addWidget(self.progress)        
        progress_group.setLayout(progress_layout)
        output_layout.addWidget(progress_group)        
        output_group.setLayout(output_layout)
        self.layout.addWidget(output_group)        
        self.layout.addItem(QSpacerItem(20, 30, QSizePolicy.Minimum, QSizePolicy.Expanding))        
        button_layout = QHBoxLayout()
        button_layout.setAlignment(Qt.AlignCenter)        
        self.scrape_btn = QPushButton("START SCRAPING")
        self.scrape_btn.setFixedSize(250, 60)
        self.scrape_btn.setFont(QFont("Segoe UI", 14, QFont.Bold))
        self.scrape_btn.clicked.connect(self.start_scraping)
        button_layout.addWidget(self.scrape_btn)        
        self.layout.addLayout(button_layout)        
        footer_label = QLabel("© 2024 Ajantha Bathroom Products and Pipes Pvt.Ltd | Python Desktop Application | DTCP Plan Permit Automation")
        footer_label.setAlignment(Qt.AlignCenter)
        footer_label.setStyleSheet("color: #666666; font-size: 10px; margin-top: 20px; background: transparent; font-weight: 600;")
        self.layout.addWidget(footer_label)        
        self.setLayout(self.layout)

    def start_scraping(self):
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", f"DTCP_Results_{self.district.currentText()}_{self.year.currentText()}.xlsx", "Excel Files (*.xlsx)")        
        if not save_path:
            return            
        self.scrape_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.loader.setVisible(True)        
        if hasattr(self, 'loader_movie'):
            self.loader_movie.start()        
        self.thread = ScraperThread(
            self.department.currentText(),
            self.approval_type.currentText(),
            self.district.currentText(),
            self.year.currentText(),
            save_path
        )
        self.thread.finished_signal.connect(self.on_scraping_finished)
        self.thread.progress_signal.connect(self.update_progress)
        self.thread.start() 
    
    def update_progress(self, current, total):
        self.progress.setMaximum(total)
        self.progress.setValue(current)

    def on_scraping_finished(self, message):
        self.scrape_btn.setEnabled(True)
        self.progress.setVisible(False)
        self.loader.setVisible(False)        
        if hasattr(self, 'loader_movie'):
            self.loader_movie.stop()       
        success_msg = f"""
        <div style='font-family: Segoe UI; font-size: 14px; color: #ffffff; background: #000000; padding: 25px; border-radius: 8px; border: 2px solid #dc2626; max-width: 500px;'>
            <div style='text-align: center; margin-bottom: 20px;'>
                <h3 style='color: #dc2626; margin: 0; font-size: 22px; font-weight: 800;'>✅ SCRAPING COMPLETED</h3>
                <div style='color: #cccccc; font-size: 13px; margin-top: 5px;'>DTCP data successfully extracted</div>
            </div>
            
            <div style='background: #1a1a1a; padding: 15px; border-radius: 6px; margin-bottom: 15px;'>
                <div style='display: flex; justify-content: space-between;'>
                    <div style='text-align: center; flex: 1;'>
                        <div style='font-size: 20px; font-weight: 800; color: #dc2626;'>10</div>
                        <div style='font-size: 11px; color: #cccccc;'>Documents Processed</div>
                    </div>
                    <div style='text-align: center; flex: 1;'>
                        <div style='font-size: 20px; font-weight: 800; color: #dc2626;'>{self.district.currentText()}</div>
                        <div style='font-size: 11px; color: #cccccc;'>District</div>
                    </div>
                    <div style='text-align: center; flex: 1;'>
                        <div style='font-size: 20px; font-weight: 800; color: #dc2626;'>{self.year.currentText()}</div>
                        <div style='font-size: 11px; color: #cccccc;'>Year</div>
                    </div>
                </div>
            </div>
            
            <div style='margin-top: 15px; padding: 12px; background: #dc2626; border-radius: 4px; text-align: center;'>
                <span style='color: #ffffff; font-weight: 700; font-size: 13px;'>📊 DATA READY FOR ANALYSIS</span>
            </div>
        </div>
        """        
        msg_box = QMessageBox()
        msg_box.setWindowTitle("DTCP Scraping Completed")
        msg_box.setTextFormat(Qt.RichText)
        msg_box.setText(success_msg)
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()

if __name__ == "__main__":
    app = QApplication(sys.argv)    
    app.setApplicationName("DTCP Plan Permit Scraper")
    app.setApplicationVersion("1.0")
    app.setOrganizationName("Ajantha Bathroom Products")
    app.setOrganizationDomain("ajantha.com")   
    win = DTCPApp()
    win.show()
    sys.exit(app.exec_())